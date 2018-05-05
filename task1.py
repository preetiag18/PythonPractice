import openpyxl
import os
def main():
	path = os.getcwd()
	#print(path)
	files = os.listdir(path)
	files_xlsx = [f for f in files if f[-4:] =='xlsx']
	#print(files_xlsx)
	for file in files_xlsx:
		#File to be copied
		wb = openpyxl.load_workbook(file)#Add file name
		sheet = wb["Sheet1"]#Add Sheet name
		#File to be pasted into
		template = openpyxl.load_workbook(path+"\\result\SoftDataUpload.xlsx")#Add file name
		temp_sheet = template["SampleSoftData"] #Add Sheet name
		maxRow = temp_sheet.max_row
		temp_sheet['B'+str(maxRow+1)].value = sheet['B11'].value
		temp_sheet['B'+str(maxRow+2)].value = sheet['N11'].value
		temp_sheet['D'+str(maxRow+1)].value = sheet['C14'].value
		temp_sheet['D'+str(maxRow+2)].value = sheet['O14'].value
		temp_sheet['E'+str(maxRow+1)].value = sheet['C15'].value
		temp_sheet['E'+str(maxRow+2)].value = sheet['O15'].value
		temp_sheet['H'+str(maxRow+1)].value = sheet['D17'].value
		temp_sheet['H'+str(maxRow+2)].value = sheet['P17'].value 
		temp_sheet['I'+str(maxRow+1)].value = sheet['C16'].value.strip('Mobile :')
		temp_sheet['I'+str(maxRow+2)].value = sheet['O16'].value.strip('Mobile :')
		temp_sheet['J'+str(maxRow+1)].value = sheet['B23'].value
		temp_sheet['J'+str(maxRow+2)].value = sheet['N23'].value
		temp_sheet['K'+str(maxRow+1)].value = sheet['G23'].value
		temp_sheet['K'+str(maxRow+2)].value = sheet['S23'].value
		temp_sheet['M'+str(maxRow+1)].value = sheet['J29'].value
		temp_sheet['M'+str(maxRow+2)].value = sheet['V29'].value
		temp_sheet['N'+str(maxRow+1)].value = sheet['B8'].value
		temp_sheet['N'+str(maxRow+2)].value = sheet['N8'].value
		temp_sheet['O'+str(maxRow+1)].value = sheet['I8'].value
		temp_sheet['O'+str(maxRow+2)].value = sheet['U8'].value
		temp_sheet['Q'+str(maxRow+1)].value = sheet['I28'].value
		temp_sheet['Q'+str(maxRow+2)].value = sheet['U28'].value
		temp_sheet['R'+str(maxRow+1)].value = sheet['J23'].value
		temp_sheet['R'+str(maxRow+2)].value = sheet['V23'].value
		temp_sheet['S'+str(maxRow+1)].value = sheet['J28'].value
		temp_sheet['S'+str(maxRow+2)].value = sheet['V28'].value
		temp_sheet['P'+str(maxRow+1)].value = sheet['B5'].value.strip('/UIN- ')
		temp_sheet['P'+str(maxRow+2)].value = sheet['N5'].value.strip('/UIN- ')
		temp_sheet['A'+str(maxRow+1)].value = "AWBPPD001"
		temp_sheet['A'+str(maxRow+2)].value = "AWBPPD001"
		city1 = sheet['C15'].value.split(",")
		city2 = sheet['O15'].value.split(",")
		temp_sheet['F'+str(maxRow+1)].value = city1[len(city1)-1]
		temp_sheet['F'+str(maxRow+2)].value = city2[len(city2)-1] 
		
		
		template.save(path+"\\result\SoftDataUpload.xlsx")
if __name__ == '__main__':
    main()